import os
import openpyxl
from backend.db.database import SessionLocal, init_db, engine
from backend.db.models import Base, Account, Order, Ticket, CustomerAgreement, KnownIssue

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "AI Agent Assessment - Candidate Pack", "ParcelPilot_Assessment_Data.xlsx")

def seed_database():
    print("Initializing database tables...")
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)

    session = SessionLocal()

    try:
        print(f"Loading Excel data from: {EXCEL_PATH}")
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

        # 1. Seed Accounts
        ws_accounts = wb["accounts"]
        account_rows = list(ws_accounts.iter_rows(values_only=True))
        account_headers = [str(h).strip().lower() for h in account_rows[0] if h is not None]
        print(f"Seeding {len(account_rows) - 1} accounts...")

        for row in account_rows[1:]:
            if not row or not row[0]:
                continue
            acc_id = str(row[0]).strip()
            acc_name = str(row[1]).strip()
            plan = str(row[2]).strip()
            status = str(row[3]).strip()
            csm = str(row[4]).strip() if row[4] else None
            contract_file = str(row[5]).strip() if row[5] else None
            premium_support = bool(row[6]) if len(row) > 6 and row[6] is not None else False
            notes = str(row[7]).strip() if len(row) > 7 and row[7] else None

            account = Account(
                account_id=acc_id,
                account_name=acc_name,
                plan=plan,
                status=status,
                csm=csm,
                contract_file=contract_file,
                premium_support=premium_support,
                notes=notes
            )
            session.add(account)

        # 2. Seed Customer Agreements
        print("Seeding customer agreements...")
        northstar_agreement = CustomerAgreement(
            agreement_id="AGR-001",
            account_id="ACCT-001",
            term_start="2026-01-01",
            term_end="2026-12-31",
            status="ACTIVE",
            p1_sla="15 minutes, 24x7",
            p2_sla="1 hour",
            p3_sla="8 business hours",
            free_cancellation_pre_pickup=True,
            credit_fixed_amount=None,
            credit_delay_hours=2.0,
            monthly_credit_cap=5000.0,
            raw_contract_file="05_Northstar_Logistics_Enterprise_Agreement.pdf"
        )
        session.add(northstar_agreement)

        lumenworks_agreement = CustomerAgreement(
            agreement_id="AGR-002",
            account_id="ACCT-002",
            term_start="2026-03-01",
            term_end="2027-02-28",
            status="ACTIVE",
            p1_sla="2 business hours",
            p2_sla="4 business hours",
            p3_sla="2 business days",
            free_cancellation_pre_pickup=False,
            credit_fixed_amount=300.0,
            credit_delay_hours=4.0,
            monthly_credit_cap=None,
            raw_contract_file="06_LumenWorks_Service_Agreement.pdf"
        )
        session.add(lumenworks_agreement)

        # 3. Seed Orders
        ws_orders = wb["orders"]
        order_rows = list(ws_orders.iter_rows(values_only=True))
        print(f"Seeding {len(order_rows) - 1} orders...")

        for row in order_rows[1:]:
            if not row or not row[0]:
                continue
            order_id = str(row[0]).strip()
            account_id = str(row[1]).strip()
            carrier = str(row[2]).strip()
            status = str(row[3]).strip()
            booked_at = str(row[4]).strip()
            pickup_window_start = str(row[5]).strip() if row[5] else None
            pickup_window_end = str(row[6]).strip() if row[6] else None
            pickup_actual_at = str(row[7]).strip() if row[7] else None
            shipment_fee_inr = float(row[8]) if row[8] is not None else 0.0
            carrier_fault = bool(row[9]) if row[9] is not None else False
            customer_fault = bool(row[10]) if row[10] is not None else False
            cancellation_requested_at = str(row[11]).strip() if row[11] else None
            notes = str(row[12]).strip() if len(row) > 12 and row[12] else None

            order = Order(
                order_id=order_id,
                account_id=account_id,
                carrier=carrier,
                status=status,
                booked_at=booked_at,
                pickup_window_start=pickup_window_start,
                pickup_window_end=pickup_window_end,
                pickup_actual_at=pickup_actual_at,
                shipment_fee_inr=shipment_fee_inr,
                carrier_fault=carrier_fault,
                customer_fault=customer_fault,
                cancellation_requested_at=cancellation_requested_at,
                notes=notes
            )
            session.add(order)

        # 4. Seed Tickets
        ws_tickets = wb["tickets"]
        ticket_rows = list(ws_tickets.iter_rows(values_only=True))
        print(f"Seeding {len(ticket_rows) - 1} tickets...")

        for row in ticket_rows[1:]:
            if not row or not row[0]:
                continue
            ticket_id = str(row[0]).strip()
            account_id = str(row[1]).strip()
            created_at = str(row[2]).strip()
            status = str(row[3]).strip()
            subject = str(row[4]).strip()
            description = str(row[5]).strip()
            channel = str(row[6]).strip()
            assigned_to = str(row[7]).strip() if row[7] else None
            last_customer_message_at = str(row[8]).strip() if row[8] else None
            historical_resolution = str(row[9]).strip() if len(row) > 9 and row[9] else None

            # Calculate initial severity tag
            calc_severity = "P3"
            if "outage" in subject.lower() or "failing" in subject.lower() or "exposure" in subject.lower():
                calc_severity = "P1"
            elif "bulk" in subject.lower() or "fails" in subject.lower() or "degraded" in subject.lower():
                calc_severity = "P2"

            ticket = Ticket(
                ticket_id=ticket_id,
                account_id=account_id,
                created_at=created_at,
                status=status,
                subject=subject,
                description=description,
                channel=channel,
                assigned_to=assigned_to,
                last_customer_message_at=last_customer_message_at,
                historical_resolution=historical_resolution,
                calculated_severity=calc_severity,
                is_sla_breached=False
            )
            session.add(ticket)

        # 5. Seed Known Issues from Operations Guide
        print("Seeding known issues...")
        ki_list = [
            KnownIssue(
                issue_id="KI-208",
                title="Bulk Upload failures on large CSVs",
                status="Investigating",
                opened_at="2026-08-10",
                resolved_at=None,
                description="Some Growth and Enterprise customers experience intermittent failures on CSV uploads above approximately 3,000 rows, even though the supported product limit remains 5,000 rows. Individual shipment creation is unaffected.",
                affected_plans="Growth, Enterprise",
                workaround="Split the upload into files below 3,000 rows."
            ),
            KnownIssue(
                issue_id="KI-211",
                title="SwiftShip pickup webhook delay",
                status="Monitoring",
                opened_at="2026-08-12",
                resolved_at=None,
                description="SwiftShip pickup confirmation webhooks can arrive up to 20 minutes late. A parcel may physically be collected while ParcelPilot still shows BOOKED. Before telling a customer that a pickup did not occur, verify the carrier status or wait through the known delay window.",
                affected_plans="All customers using SwiftShip",
                workaround="Verify the carrier status or wait through the 20-minute known delay window before reporting pickup failure."
            ),
            KnownIssue(
                issue_id="KI-176",
                title="Address validation",
                status="Resolved",
                opened_at="2026-07-01",
                resolved_at="2026-07-18",
                description="Address validation errors caused by special character escaping. Resolved on 18 July 2026. Do not use this resolved issue to explain new incidents unless evidence specifically matches it.",
                affected_plans="All plans",
                workaround="None required (Resolved)."
            )
        ]
        for ki in ki_list:
            session.add(ki)

        session.commit()
        print("Database successfully seeded with accounts, agreements, orders, tickets, and known issues!")

    except Exception as e:
        session.rollback()
        print(f"Error seeding database: {e}")
        raise
    finally:
        session.close()

if __name__ == "__main__":
    seed_database()
